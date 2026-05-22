"""
AutoScout24 Agent — v7 (2026-05-21)
Versión GitHub Actions
- Config en config.json (múltiples búsquedas, filtros opcionales)
- Credenciales Google: variable GOOGLE_CREDENTIALS_JSON (base64)
- Contraseña email:    variable EMAIL_PASSWORD
- Estado entre ejecuciones: Google Sheets (pestaña _Estado_)
- Scraping: extracción desde JSON de Next.js + fallback CSS mejorado
"""

import os, sys, json, time, base64, re, smtplib, logging, tempfile, csv, io, unicodedata
import requests
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from datetime import datetime, date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from bs4 import BeautifulSoup
import gspread
from google.oauth2.service_account import Credentials


# ══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def ruta(f): return os.path.join(SCRIPT_DIR, f)

def cargar_config() -> dict:
    path = ruta("config.json")
    if not os.path.exists(path):
        print(f"[FATAL] No se encuentra config.json en {path}"); sys.exit(1)
    with open(path, encoding="utf-8") as f:
        cfg = json.load(f)
    cfg["email"]["password"] = os.environ.get("EMAIL_PASSWORD", "")
    if not cfg["email"]["password"]:
        print("[FATAL] Variable de entorno EMAIL_PASSWORD no definida."); sys.exit(1)
    return cfg

CFG = cargar_config()


# ══════════════════════════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════════════════════════

NIVEL = getattr(logging, CFG["log"].get("nivel", "INFO").upper(), logging.INFO)

# Buffer en memoria para capturar el log del día y publicarlo en gh-pages
_log_lines: list[str] = []

class _BufHandler(logging.Handler):
    def emit(self, record):
        _log_lines.append(self.format(record))

_LOG_FMT = "%(asctime)s [%(levelname)-8s] %(funcName)-28s — %(message)s"
logging.basicConfig(
    level=logging.DEBUG,
    format=_LOG_FMT,
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger("autoscout")
_bh = _BufHandler()
_bh.setFormatter(logging.Formatter(_LOG_FMT, datefmt="%Y-%m-%d %H:%M:%S"))
_bh.setLevel(logging.DEBUG)
log.addHandler(_bh)


# ══════════════════════════════════════════════════════════════
#  CREDENCIALES GOOGLE
# ══════════════════════════════════════════════════════════════

_creds_file = None

def obtener_credenciales_google() -> str:
    global _creds_file
    if _creds_file and os.path.exists(_creds_file):
        return _creds_file
    b64 = os.environ.get("GOOGLE_CREDENTIALS_JSON", "")
    if not b64:
        log.critical("Variable GOOGLE_CREDENTIALS_JSON no definida."); sys.exit(1)
    try:
        json_bytes = base64.b64decode(b64)
        json.loads(json_bytes)
    except Exception as e:
        log.critical(f"GOOGLE_CREDENTIALS_JSON inválido: {e}"); sys.exit(1)
    tmp = tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="wb")
    tmp.write(json_bytes); tmp.close()
    _creds_file = tmp.name
    log.debug(f"Credenciales Google en temporal: {_creds_file}")
    return _creds_file

def limpiar_temporales():
    global _creds_file
    if _creds_file and os.path.exists(_creds_file):
        try: os.unlink(_creds_file)
        except: pass


# ══════════════════════════════════════════════════════════════
#  SCRAPING — HELPERS DE PARSEO DE CAMPOS
# ══════════════════════════════════════════════════════════════

def limpiar_numero(texto: str) -> int:
    """Convierte '21.350 €' o '45.900 km' en 21350 / 45900."""
    nums = re.sub(r"[^\d]", "", str(texto))
    return int(nums) if nums else 0

def es_precio_razonable(n: int) -> bool:
    return 500 < n < 500_000

def es_km_razonable(n: int) -> bool:
    return 0 <= n < 2_000_000


# ══════════════════════════════════════════════════════════════
#  SCRAPING — EXTRACCIÓN DESDE __NEXT_DATA__ (método principal)
# ══════════════════════════════════════════════════════════════

def _recopilar_listings_del_json(data: dict) -> list[dict]:
    """
    Recorre todo el JSON de Next.js buscando objetos que parezcan anuncios.
    Maneja el patrón Redux normalizado (IDs en array + datos en byId/dict).
    """
    encontrados = []

    def _buscar(obj, depth=0):
        if depth > 15:
            return
        if isinstance(obj, dict):
            tiene_precio   = any(k in obj for k in ("price", "pricing", "prices"))
            tiene_id       = any(k in obj for k in ("id", "guid", "listingId"))
            tiene_vehiculo = any(k in obj for k in
                                 ("vehicle", "vehicleDetails", "make", "model",
                                  "firstRegistration", "mileage"))
            if tiene_precio and tiene_id and tiene_vehiculo:
                encontrados.append(obj)
                return  # no buscar dentro para evitar duplicados
            for v in obj.values():
                if isinstance(v, (dict, list)):
                    _buscar(v, depth + 1)
        elif isinstance(obj, list):
            for item in obj:
                if isinstance(item, (dict, list)):
                    _buscar(item, depth + 1)

    _buscar(data)
    return encontrados


def _safe_dict(val) -> dict:
    """Devuelve val si es dict, o {} en caso contrario."""
    return val if isinstance(val, dict) else {}

def _safe_str(val) -> str:
    """Devuelve val como string si no es dict/list, o '' en caso contrario."""
    if val is None or isinstance(val, (dict, list)):
        return ""
    return str(val).strip()

def _parsear_listing_json(item, pais: str) -> dict | None:
    """Convierte un anuncio del JSON de Next.js en nuestro diccionario."""
    if not isinstance(item, dict):
        return None
    try:
        aid = _safe_str(item.get("id") or item.get("guid") or item.get("listingId") or "")
        if not aid:
            return None

        # URL
        url_raw = _safe_str(
            item.get("url") or item.get("link") or item.get("detailPageUrl") or
            item.get("absoluteUrl") or item.get("shareUrl") or item.get("href") or ""
        )
        if url_raw.startswith("http"):
            url = url_raw
        elif url_raw.startswith("/"):
            url = f"https://www.autoscout24.{pais}{url_raw}"
        else:
            url = ""

        # Título — evitar duplicar partes del modelo en la descripción
        titulo = ""
        for key in ("title", "name", "headline"):
            val = item.get(key)
            if val and isinstance(val, str):
                titulo = val.strip(); break
        if not titulo:
            v    = _safe_dict(item.get("vehicle"))
            make = _safe_dict(v.get("make")).get("name", "") or _safe_str(v.get("make"))
            model= _safe_dict(v.get("model")).get("name", "") or _safe_str(v.get("model"))
            desc = _safe_str(v.get("description") or v.get("modelVersionInput") or
                             v.get("version") or "")
            # Evitar duplicar si desc ya empieza con lo que hay en model
            if model and desc and desc.lower().startswith(model.lower().split()[-1].lower()):
                titulo = f"{make} {model} {desc[len(model.split()[-1]):].strip()}".strip()
            else:
                titulo = " ".join(filter(None, [make, model, desc])).strip()
            titulo = re.sub(r'\s+', ' ', titulo) or "Sin título"

        # Precio
        precio = 0
        p = item.get("price") or item.get("pricing") or {}
        if isinstance(p, dict):
            # 1) campos numéricos directos
            raw = p.get("value") or p.get("amount") or p.get("gross") or 0
            if raw:
                try: precio = int(raw)
                except: precio = limpiar_numero(str(raw))
            # 2) priceFormatted: "€ 28.899" o "28.899 €"
            if not es_precio_razonable(precio):
                pf = _safe_str(p.get("priceFormatted", ""))
                if pf:
                    precio = limpiar_numero(pf)
        elif isinstance(p, (int, float)):
            precio = int(p)
        # 3) fallback: tracking.price (siempre es numérico como string)
        if not es_precio_razonable(precio):
            tracking = _safe_dict(item.get("tracking"))
            pt = _safe_str(tracking.get("price", ""))
            if pt:
                try: precio = int(pt)
                except: precio = limpiar_numero(pt)
        if not es_precio_razonable(precio):
            precio = 0

        # vehicleDetails puede ser lista [{"iconName":..,"data":..}] o dict
        vd_raw = item.get("vehicleDetails")
        vd_list = vd_raw if isinstance(vd_raw, list) else []
        vd_icon = {d.get("iconName",""): d.get("data","") for d in vd_list if isinstance(d, dict)}
        vd   = _safe_dict(vd_raw)
        v_alt= _safe_dict(item.get("vehicle"))
        tracking_d = _safe_dict(item.get("tracking"))

        # Año — orden de prioridad: tracking.firstRegistration > vehicleDetails > vehicle
        anio = ""
        # 1. tracking.firstRegistration: "MM-YYYY" o "YYYY"
        fr_tracking = _safe_str(tracking_d.get("firstRegistration", ""))
        if fr_tracking:
            m_yr = re.search(r"(19|20)\d{2}", fr_tracking)
            if m_yr: anio = m_yr.group()
        # 2. vehicleDetails lista — iconName "calendar": "04/2019"
        if not anio:
            cal_str = vd_icon.get("calendar", "")
            if cal_str:
                m_yr = re.search(r"(19|20)\d{2}", cal_str)
                if m_yr: anio = m_yr.group()
        # 3. vehicle dict
        if not anio:
            for key in ("firstRegistration", "firstRegistrationDate", "yearOfProduction", "year"):
                val = _safe_str(v_alt.get(key) or vd.get(key) or item.get(key) or "")
                if val:
                    m_yr = re.search(r"(19|20)\d{2}", val)
                    if m_yr: anio = m_yr.group(); break

        # Km — orden de prioridad: vehicle.mileageInKm > vehicleDetails lista > tracking.mileage
        km = ""
        # 1. vehicle.mileageInKm: "96.217 km"
        mik = _safe_str(v_alt.get("mileageInKm", ""))
        if mik:
            km_val = limpiar_numero(mik)
            if es_km_razonable(km_val) and km_val > 0:
                km = f"{km_val:,} km".replace(",", ".")
        # 2. vehicleDetails lista — iconName "mileage_odometer"
        if not km:
            od_str = vd_icon.get("mileage_odometer", "")
            if od_str:
                km_val = limpiar_numero(od_str)
                if es_km_razonable(km_val) and km_val > 0:
                    km = f"{km_val:,} km".replace(",", ".")
        # 3. tracking.mileage (número puro)
        if not km:
            km_t = _safe_str(tracking_d.get("mileage", ""))
            if km_t:
                km_val = limpiar_numero(km_t)
                if es_km_razonable(km_val) and km_val > 0:
                    km = f"{km_val:,} km".replace(",", ".")

        # Combustible
        combustible = ""
        for key in ("fuelType", "fuel", "fuelTypeDetails", "fuelCategory"):
            val = vd.get(key) or v_alt.get(key) or item.get(key) or ""
            if isinstance(val, dict):
                val = val.get("name") or val.get("value") or val.get("key") or ""
            val = _safe_str(val)
            if val and len(val) < 40:
                combustible = val; break

        # Ubicación
        ubicacion = ""
        seller = _safe_dict(item.get("seller"))
        loc    = _safe_dict(seller.get("location") or item.get("location"))
        for key in ("city", "zip", "region", "countryName", "address"):
            val = _safe_str(loc.get(key) or "")
            if val:
                ubicacion = val; break

        if not aid or precio == 0:
            return None

        return {
            "id":              aid,
            "titulo":          titulo,
            "precio":          precio,
            "anio":            anio,
            "km":              km,
            "combustible":     combustible,
            "ubicacion":       ubicacion,
            "url":             url,
            "fecha_detectado": date.today().isoformat(),
        }
    except Exception as e:
        log.debug(f"Error parseando JSON item: {e}")
        return None


def extraer_desde_next_data(soup, pais: str) -> list[dict]:
    """Extrae anuncios del JSON embebido de Next.js (__NEXT_DATA__)."""
    script = soup.find("script", {"id": "__NEXT_DATA__"})
    if not script or not script.string:
        return []
    try:
        data = json.loads(script.string)
    except json.JSONDecodeError:
        return []

    listings_raw = _recopilar_listings_del_json(data)
    if not listings_raw:
        log.debug("No se encontraron listings en __NEXT_DATA__")
        return []

    anuncios = []
    for item in listings_raw:
        a = _parsear_listing_json(item, pais)
        if a:
            anuncios.append(a)

    log.debug(f"__NEXT_DATA__: {len(anuncios)} anuncios parseados de {len(listings_raw)} objetos")
    return anuncios


# ══════════════════════════════════════════════════════════════
#  SCRAPING — FALLBACK CSS MEJORADO
# ══════════════════════════════════════════════════════════════

def _extraer_precio_css(card) -> int:
    """Busca el precio evitando confundirlo con km u otros números."""
    # 1. Selectores específicos por prioridad
    selectores = [
        "[data-testid='price']",
        "[data-testid='regular-price']",
        ".cldt-price",
        "[class*='Price_price']",
        "[class*='price__']",
        "strong[class*='Price']",
        "p[class*='Price']",
    ]
    for sel in selectores:
        el = card.select_one(sel)
        if el:
            txt = el.get_text(strip=True)
            if "€" in txt:
                n = limpiar_numero(txt)
                if es_precio_razonable(n):
                    return n

    # 2. Buscar texto con € filtrando km/h y valores raros
    for el in card.find_all(string=re.compile(r"[\d\.]+\s*€")):
        txt = str(el).strip()
        if "km" in txt.lower() or "km/h" in txt.lower():
            continue
        n = limpiar_numero(txt)
        if es_precio_razonable(n):
            return n

    return 0


def _extraer_km_css(card) -> str:
    """Extrae los km del anuncio."""
    selectores = [
        "[data-testid='mileage']",
        "[data-testid='km']",
        "[class*='mileage']",
        "[class*='Mileage']",
    ]
    for sel in selectores:
        el = card.select_one(sel)
        if el:
            txt = el.get_text(strip=True)
            n = limpiar_numero(txt)
            if es_km_razonable(n) and n > 0:
                return f"{n:,} km".replace(",", ".")

    for el in card.find_all(string=re.compile(r"[\d\.]+\s*km", re.IGNORECASE)):
        txt = str(el).strip()
        if "km/h" in txt.lower():
            continue
        n = limpiar_numero(txt)
        if es_km_razonable(n) and n > 0:
            return f"{n:,} km".replace(",", ".")

    return ""


def _extraer_anio_css(card) -> str:
    """Extrae el año de primera matriculación."""
    selectores = [
        "[data-testid='first-registration']",
        "[data-testid='year']",
        "[class*='firstReg']",
        "[class*='FirstReg']",
    ]
    for sel in selectores:
        el = card.select_one(sel)
        if el:
            m = re.search(r"(19|20)\d{2}", el.get_text())
            if m: return m.group()

    for el in card.find_all(string=re.compile(r"\b(19|20)\d{2}\b")):
        m = re.search(r"\b(20[012]\d|199\d)\b", str(el))
        if m: return m.group()

    return ""



# Tipos de combustible conocidos en ES / DE / EN
_COMBUSTIBLES = {
    "gasolina", "diésel", "diesel", "híbrido", "híbrida",
    "eléctrico", "eléctrica", "glp", "gnc", "gas natural",
    "hidrógeno", "gasolina/eléctrico", "diésel/eléctrico",
    "gasolina/gas natural", "gas licuado (glp)",
    # Alemán
    "benzin", "elektro", "hybrid", "erdgas", "autogas",
    "wasserstoff", "plug-in-hybrid", "mild-hybrid",
    # Inglés
    "petrol", "electric", "lpg", "cng", "hydrogen",
}

def _extraer_combustible_css(card) -> str:
    # 1. Selectores data-testid específicos
    for sel in ["[data-testid='fuel-type']", "[data-testid='fuel']",
                "[data-testid='fuelType']"]:
        el = card.select_one(sel)
        if el:
            txt = el.get_text(strip=True)
            if txt:
                return txt

    # 2. Buscar cualquier texto que coincida con tipos conocidos
    for el in card.find_all(string=True):
        txt = el.strip()
        if not txt or len(txt) > 40:
            continue
        if txt.lower() in _COMBUSTIBLES:
            return txt
        # Coincidencia parcial para tipos compuestos
        for tipo in _COMBUSTIBLES:
            if len(tipo) > 4 and tipo in txt.lower():
                return txt

    return ""


def _extraer_ubicacion_css(card) -> str:
    # 1. Selectores data-testid específicos
    for sel in ["[data-testid='location']", "[data-testid='seller-location']",
                "[data-testid='city']", ".cldt-summary-seller-contact-location"]:
        el = card.select_one(sel)
        if el:
            txt = el.get_text(strip=True)
            if txt and len(txt) < 60:
                return txt

    # 2. Clases con "location" o "Location" (sin importar el prefijo hash)
    for el in card.find_all(attrs={"class": True}):
        clases = " ".join(el.get("class", []))
        if "ocation" in clases or "seller" in clases.lower():
            txt = el.get_text(strip=True)
            # Descartar textos que sean precio, km, año u otros campos
            if (txt and 2 < len(txt) < 60
                    and "€" not in txt
                    and "km" not in txt.lower()
                    and not re.match(r"^\d{4}$", txt)
                    and not re.match(r"^[\d\.]+$", txt)):
                return txt

    # 3. Buscar elemento que contenga un SVG de pin de mapa (📍)
    #    AutoScout24 suele usar un SVG justo antes del texto de ciudad
    for svg in card.find_all("svg"):
        siguiente = svg.find_next_sibling(string=True)
        if siguiente:
            txt = siguiente.strip()
            if txt and len(txt) < 60 and "€" not in txt and "km" not in txt.lower():
                return txt
        parent = svg.parent
        if parent:
            txt = parent.get_text(strip=True)
            if (txt and len(txt) < 60 and "€" not in txt
                    and "km" not in txt.lower()
                    and not re.match(r"^\d+$", txt)):
                return txt

    return ""


def _extraer_url_css(card, pais: str) -> str:
    """Devuelve la URL del anuncio: el primer enlace del card que no sea de dealer/filtro."""
    base    = f"https://www.autoscout24.{pais}"
    EXCLUIR = ["haendler", "dealer", "concessionnaire", "rivenditore",
               "vendeur", "javascript:", "?sort=", "?atype=", "/lst/"]

    for a in card.find_all("a", href=True):
        href = a.get("href", "").strip()
        if not href or len(href) < 6 or href.startswith("#"):
            continue
        if any(e in href for e in EXCLUIR):
            continue
        return href if href.startswith("http") else base + href

    return ""


def extraer_desde_css(soup, pais: str) -> list[dict]:
    """Extrae anuncios mediante selectores CSS como fallback."""
    cards = (
        soup.select("article[id^='listing']") or
        soup.select("article.cldt-summary-full-item") or
        soup.select("[data-testid='result-item']") or
        soup.select("article[id]")
    )
    log.debug(f"CSS fallback: {len(cards)} cards encontradas")

    anuncios = []
    for card in cards:
        try:
            aid = card.get("id", "") or card.get("data-id", "")
            if not aid:
                continue

            titulo_el = card.select_one("h2, [data-testid='title'], .cldt-summary-headline")
            if titulo_el:
                titulo = re.sub(r'\s+', ' ', titulo_el.get_text(separator=" ", strip=True))
            else:
                titulo = "Sin título"

            precio     = _extraer_precio_css(card)
            km         = _extraer_km_css(card)
            anio       = _extraer_anio_css(card)
            combustible= _extraer_combustible_css(card)
            ubicacion  = _extraer_ubicacion_css(card)
            url        = _extraer_url_css(card, pais)

            if not aid or precio == 0:
                continue

            anuncios.append({
                "id":              str(aid),
                "titulo":          titulo,
                "precio":          precio,
                "anio":            anio,
                "km":              km,
                "combustible":     combustible,
                "ubicacion":       ubicacion,
                "url":             url,
                "fecha_detectado": date.today().isoformat(),
            })
        except Exception as e:
            log.debug(f"Error CSS card: {e}")

    return anuncios


# ══════════════════════════════════════════════════════════════
#  SCRAPING — PRINCIPAL
# ══════════════════════════════════════════════════════════════

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
}


# Mapa de código de país (config "pais") → parámetro cy= de AutoScout24
_PAIS_A_CY = {
    "es": "E", "de": "D", "fr": "F", "it": "I",
    "be": "B", "nl": "NL", "at": "A", "lu": "L",
}

def construir_url(b: dict, pagina: int = 1) -> str:
    """Construye la URL de búsqueda.

    - Dominio SIEMPRE autoscout24.es (idioma castellano).
    - El país de búsqueda se controla con cy= según config "pais":
        es→E, de→D, fr→F, it→I, be→B, nl→NL, at→A, lu→L
    - Parámetros fijos: atype=C, damaged_listing=exclude, desc=0,
      powertype=kw, sort=price, ustate=N%2CU
    """
    marca  = b["marca"].lower().replace(" ", "-")
    modelo = b["modelo"].lower().replace(" ", "-")
    pais   = b.get("pais", "es").lower()
    cy     = _PAIS_A_CY.get(pais, "E")

    # Siempre autoscout24.es para obtener resultados en castellano
    base = f"https://www.autoscout24.es/lst/{marca}/{modelo}"

    params = [
        "atype=C",
        f"cy={cy}",
        "damaged_listing=exclude",
        "desc=0",
        f"page={pagina}",
        "powertype=kw",
        "sort=price",
        "ustate=N%2CU",
    ]

    def add(key, val):
        if val is not None and val != "" and val != 0:
            params.append(f"{key}={val}")

    add("fuel",      b.get("fuel"))
    add("fregfrom",  b.get("anio_min"))
    add("fregto",    b.get("anio_max"))
    add("kmfrom",    b.get("km_min"))
    add("kmto",      b.get("km_max"))
    add("pricefrom", b.get("precio_min"))
    add("priceto",   b.get("precio_max"))

    return base + "?" + "&".join(params)


def obtener_anuncios_pagina(url: str, pais: str) -> list[dict]:
    log.debug(f"GET {url}")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=25)
        log.debug(f"HTTP {resp.status_code} — {len(resp.content):,} bytes")
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")

        # Siempre usamos autoscout24.es como dominio de scraping,
        # por lo que el JSON devuelve rutas /anuncios/... → siempre base "es"
        anuncios = extraer_desde_next_data(soup, "es")
        if anuncios:
            log.info(f"  JSON: {len(anuncios)} anuncios")
            return anuncios

        # Fallback CSS
        anuncios = extraer_desde_css(soup, "es")
        log.info(f"  CSS:  {len(anuncios)} anuncios")
        return anuncios

    except requests.exceptions.Timeout:
        log.error(f"Timeout: {url}")
    except requests.exceptions.HTTPError as e:
        log.error(f"HTTP {e.response.status_code}: {url}")
    except Exception as e:
        log.error(f"Error scraping {url}: {e}", exc_info=True)
    return []


def scrape_busqueda(b: dict) -> list[dict]:
    """Ejecuta el scraping completo de una búsqueda."""
    max_pags = b.get("max_paginas", 5)
    pausa    = b.get("pausa_seg", 3)
    pais     = b.get("pais", "es")
    todos    = []

    filtros = []
    for k, label in [("precio_min","desde"), ("precio_max","hasta"), ("anio_min","año desde"),
                     ("anio_max","año hasta"), ("km_min","km desde"), ("km_max","km hasta")]:
        if b.get(k) is not None:
            filtros.append(f"{label}:{b[k]}")

    log.info(f"Buscando: {b['marca']} {b['modelo']} | {b['pais'].upper()} | {' | '.join(filtros) or 'sin filtros extra'}")

    for pagina in range(1, max_pags + 1):
        url      = construir_url(b, pagina)
        log.info(f"── Página {pagina}/{max_pags}")
        anuncios = obtener_anuncios_pagina(url, pais)
        if not anuncios:
            log.info("  Página vacía — fin del scraping")
            break
        todos.extend(anuncios)
        if pagina < max_pags:
            time.sleep(pausa)

    # Deduplicar por ID
    vistos, unicos = set(), []
    for a in todos:
        if a["id"] not in vistos:
            vistos.add(a["id"])
            unicos.append(a)

    log.info(f"Scraping completado — {len(unicos)} anuncios únicos ({len(todos)-len(unicos)} duplicados eliminados)")
    return unicos


# ══════════════════════════════════════════════════════════════
#  GOOGLE SHEETS — CONEXIÓN
# ══════════════════════════════════════════════════════════════

_spreadsheet = None

def conectar_sheets():
    global _spreadsheet
    if _spreadsheet:
        return _spreadsheet

    creds = Credentials.from_service_account_file(
        obtener_credenciales_google(),
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
    )
    client = gspread.authorize(creds)
    nombre = CFG["google_sheets"]["nombre_hoja"]

    try:
        _spreadsheet = client.open(nombre)
        log.debug(f"Hoja '{nombre}' encontrada")
    except gspread.SpreadsheetNotFound:
        _spreadsheet = client.create(nombre)
        _spreadsheet.share(CFG["email"]["destino"], perm_type="user", role="writer")
        log.info(f"Hoja '{nombre}' creada y compartida")

    return _spreadsheet


# ══════════════════════════════════════════════════════════════
#  ESTADO — GUARDADO EN GOOGLE SHEETS POR BÚSQUEDA
# ══════════════════════════════════════════════════════════════

def _nombre_hoja_estado(nombre_busqueda: str) -> str:
    safe = re.sub(r"[^\w\s-]", "", nombre_busqueda)[:20].strip()
    return f"_Estado_{safe}_"

def cargar_estado(nombre_busqueda: str) -> dict:
    try:
        sp  = conectar_sheets()
        nom = _nombre_hoja_estado(nombre_busqueda)
        try:
            ws = sp.worksheet(nom)
        except gspread.WorksheetNotFound:
            log.info(f"[{nombre_busqueda}] Sin estado previo — primera ejecución")
            return {}
        contenido = ws.acell("A1").value
        if not contenido:
            return {}
        estado = json.loads(contenido)
        log.info(f"[{nombre_busqueda}] Estado cargado — {len(estado)} anuncios conocidos")
        return estado
    except Exception as e:
        log.error(f"Error cargando estado [{nombre_busqueda}]: {e}", exc_info=True)
        return {}

def guardar_estado(nombre_busqueda: str, anuncios: list[dict]):
    """Guarda solo {id: precio} para mantenerse bien por debajo del límite de 50000 chars."""
    try:
        sp  = conectar_sheets()
        nom = _nombre_hoja_estado(nombre_busqueda)
        try:
            ws = sp.worksheet(nom)
        except gspread.WorksheetNotFound:
            ws = sp.add_worksheet(nom, rows=1, cols=1)
        # Guardar solo precio para minimizar tamaño (evita límite de 50000 chars por celda)
        estado = {a["id"]: {"precio": a["precio"]} for a in anuncios}
        json_str = json.dumps(estado, ensure_ascii=False, separators=(",", ":"))
        log.debug(f"[{nombre_busqueda}] Estado JSON: {len(json_str)} chars para {len(estado)} anuncios")
        ws.update([[json_str]], "A1")  # nuevo orden gspread 6.x
        log.info(f"[{nombre_busqueda}] Estado guardado — {len(estado)} anuncios")
    except Exception as e:
        log.error(f"Error guardando estado [{nombre_busqueda}]: {e}", exc_info=True)


# ══════════════════════════════════════════════════════════════
#  DETECCIÓN DE CAMBIOS
# ══════════════════════════════════════════════════════════════

def detectar_cambios(actuales: list[dict], anteriores: dict):
    nuevos, bajadas = [], []
    for a in actuales:
        aid = a["id"]
        if aid not in anteriores:
            nuevos.append(a)
        else:
            p_ant = anteriores[aid].get("precio", 0)
            p_act = a.get("precio", 0)
            if p_act > 0 and p_ant > 0 and p_act < p_ant:
                diferencia = p_ant - p_act
                a["precio_anterior"]   = p_ant
                a["diferencia_precio"] = diferencia
                a["porcentaje_bajada"] = round(diferencia / p_ant * 100, 1)
                bajadas.append(a)
    log.info(f"  Cambios: {len(nuevos)} nuevos | {len(bajadas)} bajadas de precio")
    return nuevos, bajadas


# ══════════════════════════════════════════════════════════════
#  GOOGLE SHEETS — ACTUALIZACIÓN DE DATOS
# ══════════════════════════════════════════════════════════════

CABECERAS = [
    "ID", "Título", "Precio (€)", "Precio Anterior (€)", "Bajada (€)", "Bajada (%)",
    "Año", "Km", "Combustible", "Ubicación", "Estado", "Fecha Detectado", "URL"
]

def _fila(a: dict, ids_nuevos: set, ids_bajadas: set) -> list:
    aid = a["id"]
    if aid in ids_bajadas:
        estado = f"⬇ BAJADA {a.get('porcentaje_bajada')}%"
    elif aid in ids_nuevos:
        estado = "🆕 NUEVO"
    else:
        estado = "Conocido"
    return [
        a.get("id",""),                a.get("titulo",""),
        a.get("precio",0),             a.get("precio_anterior",""),
        a.get("diferencia_precio",""), a.get("porcentaje_bajada",""),
        a.get("anio",""),              a.get("km",""),
        a.get("combustible",""),       a.get("ubicacion",""),
        estado,                        a.get("fecha_detectado",""),
        a.get("url",""),
    ]

def actualizar_sheets_busqueda(nombre: str, anuncios: list, nuevos: list, bajadas: list) -> str:
    try:
        sp = conectar_sheets()
        ids_nuevos  = {a["id"] for a in nuevos}
        ids_bajadas = {a["id"] for a in bajadas}

        # ── Hoja de la búsqueda (todos los anuncios) ──────────
        nombre_ws = nombre[:50]
        try:
            ws = sp.worksheet(nombre_ws)
        except gspread.WorksheetNotFound:
            ws = sp.add_worksheet(nombre_ws, rows=2000, cols=15)
            log.info(f"  Hoja '{nombre_ws}' creada")

        if not ws.row_values(1):
            ws.append_row(CABECERAS)

        ids_existentes = {
            fila[0]: i
            for i, fila in enumerate(ws.get_all_values()[1:], start=2)
            if fila
        }

        nuevas_filas = []
        batch_updates = []  # acumular actualizaciones para un solo batch_update
        for a in anuncios:
            f = _fila(a, ids_nuevos, ids_bajadas)
            if a["id"] in ids_existentes:
                n = ids_existentes[a["id"]]
                batch_updates.append({"range": f"A{n}:M{n}", "values": [f]})
            else:
                nuevas_filas.append(f)

        # Actualizar filas existentes en un solo batch (evita 429)
        if batch_updates:
            ws.batch_update(batch_updates)

        if nuevas_filas:
            ws.append_rows(nuevas_filas)
        log.info(f"  Sheets '{nombre_ws}': +{len(nuevas_filas)} nuevas, {len(batch_updates)} actualizadas")

        # ── Hoja del día ──────────────────────────────────────
        nombre_hoy = f"{date.today().isoformat()} {nombre[:20]}"
        try:
            ws_hoy = sp.worksheet(nombre_hoy)
            ws_hoy.clear()
        except gspread.WorksheetNotFound:
            ws_hoy = sp.add_worksheet(nombre_hoy, rows=200, cols=15)

        filas_hoy = [CABECERAS] + [_fila(a, ids_nuevos, ids_bajadas) for a in (nuevos + bajadas)]
        if len(filas_hoy) > 1:
            ws_hoy.update(filas_hoy, "A1")  # nuevo orden gspread 6.x
            log.info(f"  Hoja del día: {len(filas_hoy)-1} entradas")

        return sp.url

    except Exception as e:
        log.error(f"Error Sheets [{nombre}]: {e}", exc_info=True)
        return ""


# ══════════════════════════════════════════════════════════════
#  EMAIL
# ══════════════════════════════════════════════════════════════

def _tabla_anuncios_html(anuncios: list, tipo: str) -> str:
    """Genera una tabla HTML compacta con una fila por anuncio."""
    if not anuncios:
        return ""

    if tipo == "nuevo":
        cabecera_color = "#1565c0"
        titulo_seccion = f"🆕 Nuevos anuncios ({len(anuncios)})"
    else:
        cabecera_color = "#c62828"
        titulo_seccion = f"⬇️ Bajadas de precio ({len(anuncios)})"

    filas = []
    for i, a in enumerate(anuncios):
        bg = "#ffffff" if i % 2 == 0 else "#f8f9fa"
        precio_str = f"{a.get('precio', 0):,} €".replace(",", ".")
        anio   = a.get("anio", "") or "—"
        km     = a.get("km", "") or "—"
        comb   = a.get("combustible", "") or "—"
        ubic   = a.get("ubicacion", "") or "—"
        url    = a.get("url", "") or ""
        titulo = a.get("titulo", "Sin título")

        if tipo == "bajada":
            p_ant = f"{a.get('precio_anterior', 0):,} €".replace(",", ".")
            pct   = a.get("porcentaje_bajada", 0)
            precio_cell = (
                f'<span style="text-decoration:line-through;color:#999;">{p_ant}</span> '
                f'<span style="color:#c62828;font-weight:bold;">{precio_str}</span> '
                f'<span style="color:#c62828;font-size:11px;">(-{pct}%)</span>'
            )
        else:
            precio_cell = f'<span style="color:#1565c0;font-weight:bold;">{precio_str}</span>'

        link_cell = (
            f'<a href="{url}" style="background:#1565c0;color:white;padding:3px 9px;'
            f'border-radius:4px;text-decoration:none;font-size:11px;white-space:nowrap;">Ver →</a>'
        ) if url else "—"

        filas.append(f"""
        <tr style="background:{bg};border-bottom:1px solid #e0e0e0;">
            <td style="padding:7px 10px;font-size:12px;max-width:220px;">{titulo}</td>
            <td style="padding:7px 10px;font-size:12px;white-space:nowrap;">{precio_cell}</td>
            <td style="padding:7px 10px;font-size:12px;text-align:center;">{anio}</td>
            <td style="padding:7px 10px;font-size:12px;white-space:nowrap;">{km}</td>
            <td style="padding:7px 10px;font-size:12px;">{comb}</td>
            <td style="padding:7px 10px;font-size:12px;max-width:160px;">{ubic}</td>
            <td style="padding:7px 10px;text-align:center;">{link_cell}</td>
        </tr>""")

    return f"""
    <h3 style="color:{cabecera_color};font-size:14px;margin:16px 0 8px;">{titulo_seccion}</h3>
    <div style="overflow-x:auto;">
    <table style="width:100%;border-collapse:collapse;font-size:12px;
                  border:1px solid #e0e0e0;border-radius:6px;overflow:hidden;">
        <thead>
            <tr style="background:{cabecera_color};color:white;text-align:left;">
                <th style="padding:8px 10px;font-weight:600;">Título</th>
                <th style="padding:8px 10px;font-weight:600;">Precio</th>
                <th style="padding:8px 10px;font-weight:600;">Año</th>
                <th style="padding:8px 10px;font-weight:600;">Km</th>
                <th style="padding:8px 10px;font-weight:600;">Combustible</th>
                <th style="padding:8px 10px;font-weight:600;">Ubicación</th>
                <th style="padding:8px 10px;font-weight:600;">Link</th>
            </tr>
        </thead>
        <tbody>{"".join(filas)}</tbody>
    </table>
    </div>"""


def enviar_email_busqueda(nombre: str, nuevos: list, bajadas: list, url_sheets: str, filas_hist: list = None, adjuntar_hoja: bool = False):
    """Envía un email por búsqueda con el cuerpo HTML + adjunto JSON de los datos."""
    # Siempre se envía, haya cambios o no

    hoy = date.today().strftime("%d/%m/%Y")
    em  = CFG["email"]
    n_nuevos  = len(nuevos)
    n_bajadas = len(bajadas)

    # ── Cuerpo HTML ───────────────────────────────────────────
    link_sheets = (
        f'<p style="margin-top:20px;"><a href="{url_sheets}" '
        f'style="background:#388e3c;color:white;padding:8px 18px;'
        f'border-radius:4px;text-decoration:none;">📊 Ver hoja de cálculo</a></p>'
    ) if url_sheets else ""

    if n_nuevos == 0 and n_bajadas == 0:
        tablas = '''<div style="margin:24px 0;padding:18px 20px;background:#f9f9f7;border-radius:8px;
                         border-left:4px solid #d1d5db;text-align:center;">
            <p style="color:#6b7280;font-size:14px;">✅ Sin novedades en esta búsqueda hoy.</p>
            <p style="color:#9ca3af;font-size:12px;margin-top:6px;">No se han detectado anuncios nuevos ni bajadas de precio.</p>
        </div>'''
    else:
        tablas = _tabla_anuncios_html(nuevos, "nuevo") + _tabla_anuncios_html(bajadas, "bajada")

    html = f"""
    <html><body style="font-family:Arial,sans-serif;max-width:720px;margin:0 auto;padding:20px;">
        <div style="background:#1565c0;color:white;padding:16px 24px;border-radius:8px 8px 0 0;">
            <h1 style="margin:0;font-size:20px;">🚗 AutoScout24 — {nombre}</h1>
            <div style="opacity:.85;font-size:13px;margin-top:4px;">
                {hoy} &nbsp;|&nbsp; {n_nuevos} nuevos &nbsp;·&nbsp; {n_bajadas} bajadas de precio
            </div>
        </div>
        <div style="background:white;border:1px solid #e0e0e0;border-top:none;
                    padding:20px;border-radius:0 0 8px 8px;">
            {tablas}
            {link_sheets}
            <hr style="margin-top:24px;border:none;border-top:1px solid #eee;">
            <p style="color:#9e9e9e;font-size:11px;">Adjunto: datos JSON de esta búsqueda</p>
        </div>
    </body></html>"""

    # ── Adjunto JSON ──────────────────────────────────────────
    payload = {
        "busqueda":  nombre,
        "fecha":     date.today().isoformat(),
        "nuevos":    nuevos,
        "bajadas":   bajadas,
    }
    json_bytes = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    nombre_fichero = re.sub(r"[^\w\-]", "_", nombre) + f"_{date.today().isoformat()}.json"

    try:
        # Recoger todos los destinatarios válidos
        destinatarios = [em["destino"]]
        for i in range(2, 6):  # destino_2, destino_3, destino_4, destino_5
            dest = em.get(f"destino_{i}", "").strip()
            if dest and "@" in dest and "." in dest:
                destinatarios.append(dest)
        
        msg = MIMEMultipart("mixed")
        if n_nuevos == 0 and n_bajadas == 0:
            asunto = f"🚗 AutoScout24 [{nombre}] — Sin novedades — {hoy}"
        else:
            asunto = f"🚗 AutoScout24 [{nombre}] — {n_nuevos} nuevos, {n_bajadas} bajadas — {hoy}"
        msg["Subject"] = asunto
        msg["From"] = em["origen"]
        msg["To"]   = ", ".join(destinatarios)
        msg.attach(MIMEText(html, "html"))

        # Adjunto 1: JSON del día
        adjunto = MIMEApplication(json_bytes, _subtype="json")
        adjunto["Content-Disposition"] = f'attachment; filename="{nombre_fichero}"'
        msg.attach(adjunto)

        # Adjunto 2: XLSX histórico completo (si adjuntar_hoja=true en config)
        if adjuntar_hoja and filas_hist and OPENPYXL_AVAILABLE:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = nombre[:31]  # Excel limita a 31 chars
                
                # Cabeceras con estilo
                campos = list(filas_hist[0].keys()) if filas_hist else []
                ws.append(campos)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                
                # Datos
                for fila in filas_hist:
                    ws.append([fila.get(c, "") for c in campos])
                
                # Guardar en memoria
                xlsx_buf = io.BytesIO()
                wb.save(xlsx_buf)
                xlsx_bytes = xlsx_buf.getvalue()
                
                xlsx_name = re.sub(r"[^\w\-]", "_", nombre) + f"_{date.today().isoformat()}_completo.xlsx"
                adj_xlsx = MIMEApplication(xlsx_bytes, _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                adj_xlsx["Content-Disposition"] = f'attachment; filename="{xlsx_name}"'
                msg.attach(adj_xlsx)
                log.info(f"[{nombre}] XLSX histórico adjunto: {xlsx_name} ({len(filas_hist)} filas)")
            except Exception as e:
                log.error(f"Error generando XLSX: {e}")
        elif adjuntar_hoja and not OPENPYXL_AVAILABLE:
            log.warning("openpyxl no disponible - no se puede adjuntar XLSX")

        with smtplib.SMTP_SSL(em["servidor_smtp"], em["puerto_smtp"]) as server:
            server.login(em["origen"], em["password"])
            server.sendmail(em["origen"], destinatarios, msg.as_string())

        dest_str = ", ".join(destinatarios)
        log.info(f"[{nombre}] Email enviado a {len(destinatarios)} destinatario(s): {dest_str[:80]}")

    except smtplib.SMTPAuthenticationError:
        log.error("Error autenticación SMTP — usa contraseña de APLICACIÓN de Gmail (16 chars)")
    except Exception as e:
        log.error(f"Error enviando email [{nombre}]: {e}", exc_info=True)



# ══════════════════════════════════════════════════════════════
#  LECTURA COMPLETA SHEETS (histórico para dashboard)
# ══════════════════════════════════════════════════════════════

def leer_hoja_completa(nombre: str) -> list[dict]:
    """Lee TODOS los anuncios históricos de la hoja de una búsqueda desde Sheets."""
    try:
        sp  = conectar_sheets()
        ws  = sp.worksheet(nombre[:50])
        filas = ws.get_all_values()
        if not filas or len(filas) < 2:
            return []
        hdrs = filas[0]
        result = []
        for row in filas[1:]:
            if not any(row): continue
            d = {hdrs[i]: (row[i] if i < len(row) else "") for i in range(len(hdrs))}
            result.append({
                "id":                d.get("ID",""),
                "titulo":            d.get("Título",""),
                "precio":            d.get("Precio (€)",""),
                "precio_anterior":   d.get("Precio Anterior (€)",""),
                "porcentaje_bajada": d.get("Bajada (%)",""),
                "anio":              d.get("Año",""),
                "km":                d.get("Km",""),
                "combustible":       d.get("Combustible",""),
                "ubicacion":         d.get("Ubicación",""),
                "estado":            d.get("Estado",""),
                "fecha_detectado":   d.get("Fecha Detectado",""),
                "url":               d.get("URL",""),
            })
        log.info(f"[{nombre}] Histórico leído: {len(result)} anuncios de Sheets")
        return result
    except gspread.WorksheetNotFound:
        log.warning(f"[{nombre}] Hoja no encontrada para histórico")
        return []
    except Exception as e:
        log.error(f"Error leyendo histórico [{nombre}]: {e}", exc_info=True)
        return []


# ══════════════════════════════════════════════════════════════
#  GENERACIÓN HTML DASHBOARD
# ══════════════════════════════════════════════════════════════

def _slug(nombre: str) -> str:
    """'Mercedes GLC - España' → 'mercedes-glc-espana'"""
    s = unicodedata.normalize("NFD", nombre.lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-")


# ── Template HTML de página de búsqueda ───────────────────────
_HTML_BUSQUEDA = (
'<!DOCTYPE html>\n<html lang="es">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width,initial-scale=1">\n<title>__NOMBRE__ — AutoScout24 Monitor</title>\n<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600&display=swap" rel="stylesheet">\n<style>\n*{box-sizing:border-box;margin:0;padding:0}\nbody{font-family:\'IBM Plex Sans\',system-ui,sans-serif;min-height:100vh;background:var(--bg);color:var(--fg);transition:background .2s,color .2s}\n/* ── TEMAS ── */\nbody.tema-claro{--bg:#f5f5f0;--fg:#1a1a1a;--nav:#1a1a1a;--nav-fg:white;--card:white;--card-b:#ebebeb;--stat:#f9f9f7;--lbl:#aaa;--sep:#ddd;--th:#1a1a1a;--th-fg:white;--th-hov:#2d2d2d;--row-hov:#fafaf8;--td-b:#f2f2f2;--btn:#ddd;--btn-fg:#555;--inp-b:#ddd;--inp-fg:#1a1a1a;--cnt:#999;--badge-n-bg:#dbeafe;--badge-n-fg:#1e3a8a;--badge-b-bg:#fee2e2;--badge-b-fg:#7f1d1d;--badge-c-bg:#f3f4f6;--badge-c-fg:#6b7280;--pc-n:#1e40af;--pc-b:#991b1b;--pc-o:#bbb}\nbody.tema-oscuro{--bg:#0f1117;--fg:#e2e8f0;--nav:#1e2433;--nav-fg:#e2e8f0;--card:#1e2433;--card-b:#2d3748;--stat:#161b27;--lbl:#64748b;--sep:#2d3748;--th:#1e2433;--th-fg:#94a3b8;--th-hov:#2d3a52;--row-hov:#1a2236;--td-b:#2d3748;--btn:#2d3748;--btn-fg:#94a3b8;--inp-b:#2d3748;--inp-fg:#e2e8f0;--cnt:#64748b;--badge-n-bg:#1e3a5f;--badge-n-fg:#93c5fd;--badge-b-bg:#450a0a;--badge-b-fg:#fca5a5;--badge-c-bg:#1f2937;--badge-c-fg:#9ca3af;--pc-n:#60a5fa;--pc-b:#f87171;--pc-o:#475569}\nbody.tema-verde{--bg:#f0f7f0;--fg:#1a2e1a;--nav:#1a3a1a;--nav-fg:white;--card:white;--card-b:#c6e6c6;--stat:#edf7ed;--lbl:#6b8f6b;--sep:#c6e6c6;--th:#1a3a1a;--th-fg:white;--th-hov:#2d5a2d;--row-hov:#f5fbf5;--td-b:#e0f0e0;--btn:#c6e6c6;--btn-fg:#1a3a1a;--inp-b:#c6e6c6;--inp-fg:#1a2e1a;--cnt:#6b8f6b;--badge-n-bg:#dcfce7;--badge-n-fg:#14532d;--badge-b-bg:#fef9c3;--badge-b-fg:#713f12;--badge-c-bg:#f0fdf4;--badge-c-fg:#4b5563;--pc-n:#166534;--pc-b:#92400e;--pc-o:#9ca3af}\nnav{background:var(--nav);color:var(--nav-fg);padding:14px 24px;display:flex;align-items:center;gap:14px;flex-wrap:wrap}\nnav a{color:#aaa;text-decoration:none;font-size:13px}\nnav a:hover{color:var(--nav-fg)}\nnav h1{font-size:15px;font-weight:600;flex:1;min-width:180px}\n.upd{font-size:12px;color:#666}\n.tema-sel{margin-left:auto;display:flex;gap:6px;align-items:center}\n.tema-sel span{font-size:11px;color:#888}\n.tema-btn{width:18px;height:18px;border-radius:50%;border:2px solid transparent;cursor:pointer;transition:border .15s}\n.tema-btn:hover,.tema-btn.on{border-color:var(--nav-fg)}\n.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:10px;padding:16px 24px;background:var(--card);border-bottom:1px solid var(--card-b)}\n.stat{background:var(--stat);border-radius:8px;padding:10px 12px}\n.slabel{font-size:10px;color:var(--lbl);text-transform:uppercase;letter-spacing:.4px;margin-bottom:4px}\n.svalue{font-size:18px;font-weight:600}\n.sv-blue{color:var(--pc-n)}.sv-red{color:var(--pc-b)}\n.controls{padding:12px 24px;background:var(--card);border-bottom:1px solid var(--card-b);display:flex;gap:8px;flex-wrap:wrap;align-items:center}\n.controls input,.controls select{border:1px solid var(--inp-b);border-radius:6px;padding:6px 10px;font-size:13px;font-family:inherit;background:var(--card);color:var(--inp-fg)}\n.controls input:focus,.controls select:focus{outline:none;border-color:var(--fg)}\n.controls input{min-width:200px}\n.sep{color:var(--sep);font-size:12px}\n.count{padding:8px 24px 0;font-size:12px;color:var(--cnt)}\n.tw{padding:12px 24px 24px;overflow-x:auto}\ntable{width:100%;border-collapse:collapse;background:var(--card);border-radius:10px;overflow:hidden;font-size:13px;box-shadow:0 1px 4px rgba(0,0,0,.07)}\nthead{background:var(--th);color:var(--th-fg)}\nth{padding:10px 12px;text-align:left;font-weight:500;font-size:11px;white-space:nowrap;cursor:pointer;user-select:none}\nth:hover{background:var(--th-hov)}\nth.asc::after{content:" ↑"}th.desc::after{content:" ↓"}\ntd{padding:9px 12px;border-bottom:1px solid var(--td-b);vertical-align:middle}\ntr:last-child td{border:none}\ntr:hover td{background:var(--row-hov)}\n.badge{display:inline-block;font-size:11px;padding:2px 8px;border-radius:12px;font-weight:500;white-space:nowrap}\n.bn{background:var(--badge-n-bg);color:var(--badge-n-fg)}\n.bb{background:var(--badge-b-bg);color:var(--badge-b-fg)}\n.bc{background:var(--badge-c-bg);color:var(--badge-c-fg)}\n.pm{font-weight:600}.pb{color:var(--pc-n)}.pr{color:var(--pc-b)}\n.po{font-size:11px;color:var(--pc-o);text-decoration:line-through;display:block}\n.btn-ver{display:inline-block;padding:4px 10px;border:1px solid var(--btn);border-radius:5px;font-size:11px;color:var(--btn-fg);text-decoration:none;white-space:nowrap}\n.btn-ver:hover{border-color:var(--fg);color:var(--fg)}\n.pag{padding:12px 24px;display:flex;gap:5px;align-items:center;justify-content:center;flex-wrap:wrap}\n.pag button{padding:5px 11px;border:1px solid var(--btn);border-radius:5px;background:var(--card);cursor:pointer;font-size:13px;font-family:inherit;color:var(--fg)}\n.pag button:hover{border-color:var(--fg)}\n.pag button.on{background:var(--th);color:var(--th-fg);border-color:var(--th)}\n.pinfo{font-size:12px;color:var(--cnt);margin:0 6px}\n.empty{text-align:center;padding:40px;color:var(--cnt);font-size:13px}\n</style>\n</head>\n<body class="tema-claro">\n<nav>\n  <a href="../index.html">&#8592; Inicio</a>\n  <h1>__NOMBRE__</h1>\n  <span class="upd">Actualizado: __FECHA__</span>\n  <div class="tema-sel">\n    <span>Tema</span>\n    <div class="tema-btn on" id="tc" title="Claro" style="background:#f5f5f0" onclick="setTema(\'tema-claro\',this)"></div>\n    <div class="tema-btn" id="to" title="Oscuro" style="background:#0f1117" onclick="setTema(\'tema-oscuro\',this)"></div>\n    <div class="tema-btn" id="tv" title="Verde" style="background:#1a3a1a" onclick="setTema(\'tema-verde\',this)"></div>\n  </div>\n</nav>\n<div class="stats" id="stats"></div>\n<div class="controls">\n  <input type="text" id="q" placeholder="Buscar t\u00edtulo, ciudad...">\n  <select id="fe">\n    <option value="">Todos los estados</option>\n    <option value="NUEVO">Nuevos (hoy)</option>\n    <option value="BAJADA">Bajadas (hoy)</option>\n    <option value="conocido">Conocidos</option>\n  </select>\n  <span class="sep">|</span>\n  <label style="font-size:12px;color:var(--lbl)">A\u00f1o desde</label>\n  <select id="faf"><option value="">\u2014</option></select>\n  <label style="font-size:12px;color:var(--lbl)">hasta</label>\n  <select id="fat"><option value="">\u2014</option></select>\n  <span class="sep">|</span>\n  <label style="font-size:12px;color:var(--lbl)">Precio m\u00e1x</label>\n  <select id="fpm"><option value="">\u2014</option></select>\n  <span class="sep">|</span>\n  <label style="font-size:12px;color:var(--lbl)">Km m\u00e1x</label>\n  <select id="fkm"><option value="">\u2014</option></select>\n</div>\n<div class="count" id="cnt"></div>\n<div class="tw">\n<table>\n<thead><tr>\n  <th data-c="estado">Estado</th>\n  <th data-c="titulo">T\u00edtulo</th>\n  <th data-c="precio">Precio</th>\n  <th data-c="anio">A\u00f1o</th>\n  <th data-c="kmn">Km</th>\n  <th>Combustible</th>\n  <th>Ubicaci\u00f3n</th>\n  <th data-c="fecha_detectado">Detectado</th>\n  <th></th>\n</tr></thead>\n<tbody id="tb"></tbody>\n</table>\n<p class="empty" id="emp" style="display:none">Sin resultados.</p>\n</div>\n<div class="pag" id="pag"></div>\n<script>\nconst RAW=__DATA__;\nconst PG=100;\nlet sC="precio",sD="asc",pg=0;\nfunction setTema(t,el){document.body.className=t;localStorage.setItem("as24-tema",t);document.querySelectorAll(".tema-btn").forEach(b=>b.classList.remove("on"));el.classList.add("on");}\n(function(){const t=localStorage.getItem("as24-tema")||"tema-claro";const map={"tema-claro":"tc","tema-oscuro":"to","tema-verde":"tv"};document.body.className=t;const el=document.getElementById(map[t]);if(el){document.querySelectorAll(".tema-btn").forEach(b=>b.classList.remove("on"));el.classList.add("on");}})();\nfunction kmn(s){return s?parseInt(s.replace(/\\./g,"").replace(/[^\\d]/g,""))||0:0}\nfunction fp(n){n=parseInt(n)||0;return n?n.toLocaleString("es-ES")+" \u20ac":"\u2014"}\nconst ROWS=RAW.map(r=>({...r,precio:parseInt(r.precio)||0,precio_anterior:parseInt(r.precio_anterior)||0,kmn:kmn(r.km)}));\n(function init(){\n  const anios=[...new Set(ROWS.map(r=>r.anio).filter(Boolean))].sort();\n  ["faf","fat"].forEach(id=>{const s=document.getElementById(id);anios.forEach(a=>s.add(new Option(a,a)))});\n  const mx=Math.max(...ROWS.map(r=>r.precio),0);\n  const pm=document.getElementById("fpm");\n  [20000,25000,30000,35000,40000,45000,50000,60000,75000,100000].filter(p=>p<=mx+10000).forEach(p=>pm.add(new Option(p.toLocaleString("es-ES")+" \u20ac",p)));\n  const mxkm=Math.max(...ROWS.map(r=>r.kmn),0);\n  const fkm=document.getElementById("fkm");\n  [50000,75000,100000,125000,150000,175000,200000].filter(k=>k<=mxkm+25000).forEach(k=>fkm.add(new Option(k.toLocaleString("es-ES")+" km",k)));\n  document.querySelectorAll("th[data-c]").forEach(th=>th.addEventListener("click",()=>{const c=th.dataset.c;if(sC===c)sD=sD==="asc"?"desc":"asc";else{sC=c;sD="asc";}pg=0;render();}));\n  ["q","fe","faf","fat","fpm","fkm"].forEach(id=>{const el=document.getElementById(id);el.addEventListener(el.tagName==="INPUT"?"input":"change",()=>{pg=0;render();});});\n  const pr=ROWS.map(r=>r.precio).filter(Boolean);\n  const kms=ROWS.map(r=>r.kmn).filter(Boolean);\n  const nh=ROWS.filter(r=>(r.estado||"").includes("NUEVO")).length;\n  const bh=ROWS.filter(r=>(r.estado||"").includes("BAJADA")).length;\n  const fs=[...new Set(ROWS.map(r=>r.fecha_detectado).filter(Boolean))].sort();\n  const med=pr.length?Math.round(pr.reduce((a,b)=>a+b,0)/pr.length):0;\n  const km=kms.length?Math.round(kms.reduce((a,b)=>a+b,0)/kms.length):0;\n  document.getElementById("stats").innerHTML=[["Total",ROWS.length,""],["Precio m\u00edn",fp(Math.min(...pr)),""],["Precio medio",fp(med),""],["Precio m\u00e1x",fp(Math.max(...pr)),""],["Km medio",km?km.toLocaleString("es-ES")+" km":"\u2014",""],["Nuevos hoy",nh,"sv-blue"],["Bajadas hoy",bh,"sv-red"],["Desde",fs[0]||"\u2014",""]].map(([l,v,c])=>`<div class="stat"><div class="slabel">${l}</div><div class="svalue ${c}">${v}</div></div>`).join("");\n  render();\n})();\nfunction filt(){\n  const q=document.getElementById("q").value.toLowerCase();\n  const fe=document.getElementById("fe").value;\n  const af=document.getElementById("faf").value;\n  const at=document.getElementById("fat").value;\n  const pm=parseInt(document.getElementById("fpm").value)||0;\n  const km=parseInt(document.getElementById("fkm").value)||0;\n  return ROWS.filter(r=>{\n    if(q&&!(r.titulo||"").toLowerCase().includes(q)&&!(r.ubicacion||"").toLowerCase().includes(q))return false;\n    if(fe==="NUEVO"&&!(r.estado||"").includes("NUEVO"))return false;\n    if(fe==="BAJADA"&&!(r.estado||"").includes("BAJADA"))return false;\n    if(fe==="conocido"&&((r.estado||"").includes("NUEVO")||(r.estado||"").includes("BAJADA")))return false;\n    if(af&&(r.anio||"")<af)return false;\n    if(at&&(r.anio||"")>at)return false;\n    if(pm&&r.precio>pm)return false;\n    if(km&&r.kmn>km)return false;\n    return true;\n  });\n}\nfunction render(){\n  const rows=filt();\n  rows.sort((a,b)=>{\n    let av=a[sC],bv=b[sC];\n    if(["precio","kmn","precio_anterior"].includes(sC)){av=av||0;bv=bv||0;}\n    else if(sC==="anio"){av=parseInt(av)||0;bv=parseInt(bv)||0;}\n    else{av=(av||"").toLowerCase();bv=(bv||"").toLowerCase();}\n    return sD==="asc"?(av<bv?-1:av>bv?1:0):(av>bv?-1:av<bv?1:0);\n  });\n  document.querySelectorAll("th[data-c]").forEach(th=>{th.className=th.dataset.c===sC?sD:"";});\n  const tot=rows.length,pages=Math.max(1,Math.ceil(tot/PG));\n  pg=Math.min(pg,pages-1);\n  const sl=rows.slice(pg*PG,(pg+1)*PG);\n  document.getElementById("cnt").textContent=`Mostrando ${sl.length} de ${tot} anuncios (${ROWS.length} total)`;\n  document.getElementById("tb").innerHTML=sl.map(r=>{\n    const n=(r.estado||"").includes("NUEVO"),b=(r.estado||"").includes("BAJADA");\n    const bc=n?"bn":b?"bb":"bc";\n    const bt=n?"Nuevo hoy":b?`Bajada ${r.porcentaje_bajada||""}%`:"Conocido";\n    const pc=n?`<span class="pm pb">${fp(r.precio)}</span>`:b&&r.precio_anterior?`<span class="po">${fp(r.precio_anterior)}</span><span class="pm pr">${fp(r.precio)}</span>`:`<span class="pm">${fp(r.precio)}</span>`;\n    const lk=r.url?`<a href="${r.url}" target="_blank" class="btn-ver">Ver &#8594;</a>`:"\u2014";\n    return `<tr><td><span class="badge ${bc}">${bt}</span></td><td style="max-width:220px">${r.titulo||"\u2014"}</td><td style="white-space:nowrap">${pc}</td><td>${r.anio||"\u2014"}</td><td style="white-space:nowrap">${r.km||"\u2014"}</td><td style="font-size:12px;color:var(--lbl)">${r.combustible||"\u2014"}</td><td style="font-size:12px;color:var(--lbl)">${r.ubicacion||"\u2014"}</td><td style="font-size:12px;color:var(--cnt)">${r.fecha_detectado||"\u2014"}</td><td>${lk}</td></tr>`;\n  }).join("");\n  document.getElementById("emp").style.display=tot?"none":"block";\n  const pag=document.getElementById("pag");\n  if(pages<=1){pag.innerHTML="";return;}\n  let h=`<span class="pinfo">${tot} resultados</span>`;\n  const fr=Math.max(0,pg-2),to=Math.min(pages-1,pg+2);\n  if(fr>0)h+=`<button onclick="pg=0;render()">1</button>${fr>1?\'<span class="pinfo">\u2026</span>\':\'\'}`;\n  for(let i=fr;i<=to;i++)h+=`<button class="${i===pg?"on":""}" onclick="pg=${i};render()">${i+1}</button>`;\n  if(to<pages-1)h+=`${to<pages-2?\'<span class="pinfo">\u2026</span>\':\'\'}`;\n  pag.innerHTML=h;\n}\n</script>\n</body>\n</html>'
)

_HTML_INDEX = (
'<!DOCTYPE html>\n<html lang="es">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width,initial-scale=1">\n<title>AutoScout24 Monitor</title>\n<!-- v7 (2026-05-21) -->\n<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600&display=swap" rel="stylesheet">\n<style>\n*{box-sizing:border-box;margin:0;padding:0}\nbody{font-family:\'IBM Plex Sans\',system-ui,sans-serif;min-height:100vh;background:var(--bg);color:var(--fg);transition:background .2s,color .2s}\n/* ── TEMAS ── */\nbody.tema-claro{--bg:#f5f5f0;--fg:#1a1a1a;--hdr:#1a1a1a;--hdr-fg:white;--card:#fff;--card-b:#ebebeb;--stat-bg:#f9f9f7;--stat-fg:#1a1a1a;--lbl:#aaa;--val:#1a1a1a;--val-blue:#1e40af;--val-red:#991b1b}\nbody.tema-oscuro{--bg:#0f1117;--fg:#e2e8f0;--hdr:#1e2433;--hdr-fg:#e2e8f0;--card:#1e2433;--card-b:#2d3748;--stat-bg:#161b27;--stat-fg:#e2e8f0;--lbl:#64748b;--val:#e2e8f0;--val-blue:#60a5fa;--val-red:#f87171}\nbody.tema-verde{--bg:#f0f7f0;--fg:#1a2e1a;--hdr:#1a3a1a;--hdr-fg:white;--card:#fff;--card-b:#c6e6c6;--stat-bg:#edf7ed;--stat-fg:#1a2e1a;--lbl:#6b8f6b;--val:#1a2e1a;--val-blue:#166534;--val-red:#92400e}\nbody.tema-azul{--bg:#eff6ff;--fg:#1e3a8a;--hdr:#1e40af;--hdr-fg:white;--card:#fff;--card-b:#bfdbfe;--stat-bg:#dbeafe;--stat-fg:#1e3a8a;--lbl:#60a5fa;--val:#1e3a8a;--val-blue:#1d4ed8;--val-red:#dc2626}\nbody.tema-purpura{--bg:#faf5ff;--fg:#581c87;--hdr:#7c3aed;--hdr-fg:white;--card:#fff;--card-b:#e9d5ff;--stat-bg:#f3e8ff;--stat-fg:#581c87;--lbl:#a78bfa;--val:#581c87;--val-blue:#7c3aed;--val-red:#dc2626}\nheader{background:var(--hdr);color:var(--hdr-fg);padding:20px 24px;display:flex;align-items:center;gap:16px;flex-wrap:wrap}\nheader h1{font-size:18px;font-weight:600;flex:1;min-width:200px}\nheader p{font-size:13px;opacity:.7;margin-top:4px}\n.tema-sel{display:flex;gap:6px;align-items:center;margin-left:auto}\n.tema-sel span{font-size:11px;opacity:.7}\n.tema-btn{width:18px;height:18px;border-radius:50%;border:2px solid transparent;cursor:pointer;transition:border .15s}\n.tema-btn:hover,.tema-btn.on{border-color:var(--hdr-fg)}\n.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px;padding:24px}\na.card{background:var(--card);border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.07);border:1px solid var(--card-b);text-decoration:none;color:inherit;display:block;transition:box-shadow .15s,transform .15s}\na.card:hover{box-shadow:0 4px 14px rgba(0,0,0,.1);transform:translateY(-1px)}\n.card h2{font-size:14px;font-weight:600;margin-bottom:14px;color:var(--fg)}\n.cs{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:14px}\n.c{background:var(--stat-bg);border-radius:7px;padding:8px 10px}\n.cl{font-size:10px;color:var(--lbl);text-transform:uppercase;letter-spacing:.3px;margin-bottom:3px}\n.cv{font-size:17px;font-weight:600;color:var(--stat-fg)}\n.cvb{color:var(--val-blue)}.cvr{color:var(--val-red)}\n.cf{font-size:11px;color:var(--lbl);border-top:1px solid var(--card-b);padding-top:10px;margin-top:2px;display:flex;justify-content:space-between}\n.cf span{color:var(--val);font-weight:500;font-size:12px}\n</style>\n</head>\n<body class="tema-claro">\n<header>\n  <div>\n    <h1>AutoScout24 Monitor</h1>\n    <p>Actualizado: __FECHA__</p>\n  </div>\n  <div class="tema-sel">\n    <span>Tema</span>\n    <div class="tema-btn on" id="tc" title="Claro" style="background:#f5f5f0" onclick="setTema(\'tema-claro\',this)"></div>\n    <div class="tema-btn" id="to" title="Oscuro" style="background:#0f1117" onclick="setTema(\'tema-oscuro\',this)"></div>\n    <div class="tema-btn" id="tv" title="Verde" style="background:#1a3a1a" onclick="setTema(\'tema-verde\',this)"></div>\n    <div class="tema-btn" id="ta" title="Azul" style="background:#1e40af" onclick="setTema(\'tema-azul\',this)"></div>\n    <div class="tema-btn" id="tp" title="Púrpura" style="background:#7c3aed" onclick="setTema(\'tema-purpura\',this)"></div>\n  </div>\n</header>\n<div class="grid">__CARDS__</div>\n<script>\nfunction setTema(t,el){document.body.className=t;localStorage.setItem("as24-tema",t);document.querySelectorAll(".tema-btn").forEach(b=>b.classList.remove("on"));el.classList.add("on");}\n(function(){const t=localStorage.getItem("as24-tema")||"tema-claro";const map={"tema-claro":"tc","tema-oscuro":"to","tema-verde":"tv","tema-azul":"ta","tema-purpura":"tp"};document.body.className=t;const el=document.getElementById(map[t]);if(el){document.querySelectorAll(".tema-btn").forEach(b=>b.classList.remove("on"));el.classList.add("on");}})();\n</script>\n</body>\n</html>'
)



# ══════════════════════════════════════════════════════════════
#  VERIFICACIÓN ANUNCIOS ACTIVOS (limpieza de URLs muertas)
# ══════════════════════════════════════════════════════════════

def verificar_anuncios_activos(filas_hist: list[dict]) -> list[dict]:
    """Verifica cada URL y elimina anuncios que ya no están disponibles.
    Retorna la lista actualizada sin los anuncios eliminados."""
    if not filas_hist:
        return []
    
    activos = []
    eliminados = []
    
    for fila in filas_hist:
        url = fila.get("url", "")
        if not url:
            activos.append(fila)
            continue
        
        try:
            r = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
            if r.status_code == 200:
                # Verificar si el texto indica que no está disponible
                if "Este vehículo ya no está disponible" in r.text or                    "This vehicle is no longer available" in r.text or                    "Dieses Fahrzeug ist nicht mehr verfügbar" in r.text:
                    eliminados.append(fila.get("id", url[:30]))
                    log.info(f"  Anuncio eliminado (no disponible): {fila.get('titulo', '')[:50]}")
                else:
                    activos.append(fila)
            elif r.status_code == 404:
                eliminados.append(fila.get("id", url[:30]))
                log.info(f"  Anuncio eliminado (404): {fila.get('titulo', '')[:50]}")
            else:
                # Si hay error pero no es 404, lo mantenemos
                activos.append(fila)
        except Exception as e:
            # Error de red, mantener el anuncio
            activos.append(fila)
            log.debug(f"  Error verificando {url[:40]}: {e}")
        
        # Pausa breve entre requests
        time.sleep(0.5)
    
    if eliminados:
        log.info(f"Anuncios eliminados (no disponibles): {len(eliminados)}")
    
    return activos


def generar_html_busqueda(nombre: str, filas: list[dict], nuevos: list, bajadas: list, fecha: str) -> str:
    """Genera el HTML completo de una página de búsqueda con datos históricos de Sheets."""
    json_data = json.dumps(filas, ensure_ascii=False)
    html = _HTML_BUSQUEDA
    html = html.replace("__NOMBRE__", nombre)
    html = html.replace("__FECHA__", fecha)
    html = html.replace("__DATA__", json_data)
    return html


def generar_index_html(busquedas_info: list[dict]) -> str:
    """Genera el index.html con cards para cada búsqueda."""
    hoy = date.today().strftime("%d/%m/%Y")
    cards = ""
    for b in busquedas_info:
        cards += f"""<a class="card" href="busquedas/{b['slug']}.html">
  <h2>{b['nombre']}</h2>
  <div class="cs">
    <div class="c"><div class="cl">Total</div><div class="cv">{b['total']}</div></div>
    <div class="c"><div class="cl">Nuevos hoy</div><div class="cv cvb">{b['nuevos']}</div></div>
    <div class="c"><div class="cl">Bajadas hoy</div><div class="cv cvr">{b['bajadas']}</div></div>
    <div class="c"><div class="cl">Precio mín</div><div class="cv" style="font-size:14px">{b['min_precio']}</div></div>
    <div class="c"><div class="cl">Precio med</div><div class="cv" style="font-size:14px">{b['med_precio']}</div></div>
    <div class="c"><div class="cl">Última act.</div><div class="cv" style="font-size:13px">{b['fecha']}</div></div>
  </div>
  <div class="cf"><span>Ver detalle →</span><span style="color:#aaa;font-size:11px">{b['total']} anuncios históricos</span></div>
</a>"""
    html = _HTML_INDEX
    html = html.replace("__FECHA__", hoy)
    html = html.replace("__CARDS__", cards)
    return html


# ══════════════════════════════════════════════════════════════
#  PUBLICACIÓN GITHUB PAGES
# ══════════════════════════════════════════════════════════════



def generar_ver_registros() -> str:
    """Genera ver_registros.html — visor de logs.
    No enlazada desde ninguna otra página; solo accesible por URL directa."""
    return '<!DOCTYPE html>\n<html lang="es">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width,initial-scale=1">\n<title>Ver registros — AutoScout24 Monitor</title>\n<!-- v7 (2026-05-21) -->\n<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@400;500;600&display=swap" rel="stylesheet">\n<style>\n*{box-sizing:border-box;margin:0;padding:0}\nbody{font-family:\'IBM Plex Sans\',system-ui,sans-serif;background:#0f1117;color:#e2e8f0;min-height:100vh}\nheader{background:#1e2433;border-bottom:1px solid #2d3748;padding:16px 24px}\nheader h1{font-size:16px;font-weight:600}\nheader p{font-size:12px;color:#64748b;margin-top:2px}\n.main{max-width:1200px;margin:0 auto;padding:24px}\n.card{background:#1e2433;border:1px solid #2d3748;border-radius:10px;padding:20px;margin-bottom:16px}\n.card h3{font-size:13px;font-weight:600;color:#94a3b8;text-transform:uppercase;letter-spacing:.5px;margin-bottom:12px}\n.selector{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px}\nbutton{padding:6px 14px;border-radius:6px;border:1px solid #2d3748;background:#161b27;color:#94a3b8;font-size:12px;cursor:pointer;transition:all .15s}\nbutton:hover{border-color:#4b5563;color:#e2e8f0}\nbutton.active{background:#1d4ed8;border-color:#1d4ed8;color:white}\n.log-viewer{background:#0d1117;border:1px solid #2d3748;border-radius:8px;padding:16px;font-family:\'IBM Plex Mono\',monospace;font-size:12px;line-height:1.6;max-height:600px;overflow-y:auto;white-space:pre-wrap;word-break:break-all}\n.log-line{margin:2px 0}\n.log-line.error{color:#f87171}\n.log-line.warning{color:#fbbf24}\n.log-line.info{color:#94a3b8}\n.log-line.debug{color:#475569}\n.status{padding:10px 12px;border-radius:6px;font-size:12px;margin-top:12px}\n.status.ok{background:#052e16;color:#86efac;border:1px solid #14532d}\n.status.err{background:#450a0a;color:#fca5a5;border:1px solid #7f1d1d}\n.status.loading{background:#161b27;color:#64748b;border:1px solid #2d3748}\n</style>\n</head>\n<body>\n<header>\n  <h1>📋 Ver registros</h1>\n  <p>Logs de ejecuciones del AutoScout24 Monitor</p>\n</header>\n<div class="main">\n  <div class="card">\n    <h3>Seleccionar log</h3>\n    <div class="selector" id="selector">\n      <div class="status loading">Cargando logs disponibles...</div>\n    </div>\n  </div>\n  <div class="card">\n    <h3>Contenido del log</h3>\n    <div class="log-viewer" id="viewer">Selecciona un log arriba para visualizarlo</div>\n  </div>\n</div>\n<script>\nconst REPO = "antrivasclaude-glitch/autoscout-monitor";\nconst BASE_URL = `https://antrivasclaude-glitch.github.io/autoscout-monitor/`;\n\nasync function cargarLogs() {\n  try {\n    const r = await fetch(`https://api.github.com/repos/${REPO}/git/trees/gh-pages?recursive=1`);\n    if (!r.ok) throw new Error("No se pudo acceder a la API");\n    \n    const data = await r.json();\n    const logs = data.tree\n      .filter(f => f.path.startsWith("log_") && f.path.endsWith(".txt"))\n      .map(f => f.path.replace("log_", "").replace(".txt", ""))\n      .sort().reverse();\n    \n    const sel = document.getElementById("selector");\n    if (logs.length === 0) {\n      sel.innerHTML = \'<div class="status err">No hay logs disponibles</div>\';\n      return;\n    }\n    \n    sel.innerHTML = logs.map(fecha => \n      `<button onclick="cargarLog(\'${fecha}\')">${fecha}</button>`\n    ).join("");\n  } catch(e) {\n    document.getElementById("selector").innerHTML = \n      `<div class="status err">Error: ${e.message}</div>`;\n  }\n}\n\nasync function cargarLog(fecha) {\n  const viewer = document.getElementById("viewer");\n  viewer.innerHTML = \'<div class="status loading">Cargando...</div>\';\n  \n  document.querySelectorAll(".selector button").forEach(b => b.classList.remove("active"));\n  event.target.classList.add("active");\n  \n  try {\n    const r = await fetch(`${BASE_URL}log_${fecha}.txt`);\n    if (!r.ok) throw new Error(`HTTP ${r.status}`);\n    \n    const texto = await r.text();\n    const lineas = texto.split("\\n").map(linea => {\n      let clase = "info";\n      if (linea.includes("[ERROR")) clase = "error";\n      else if (linea.includes("[WARNING")) clase = "warning";\n      else if (linea.includes("[DEBUG")) clase = "debug";\n      return `<div class="log-line ${clase}">${linea}</div>`;\n    }).join("");\n    \n    viewer.innerHTML = lineas || \'<div class="status err">Log vacío</div>\';\n  } catch(e) {\n    viewer.innerHTML = `<div class="status err">Error cargando log: ${e.message}</div>`;\n  }\n}\n\nwindow.addEventListener("load", cargarLogs);\n</script>\n</body>\n</html>'


def generar_editor_configuracion() -> str:
    """Genera editar_configuracion.html — página de edición de config.json.
    No enlazada desde ninguna otra página; solo accesible por URL directa."""
    return '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Editar configuración — AutoScout24 Monitor</title>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@400;500;600&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:\'IBM Plex Sans\',system-ui,sans-serif;background:#0f1117;color:#e2e8f0;min-height:100vh}
header{background:#1e2433;border-bottom:1px solid #2d3748;padding:16px 24px;display:flex;align-items:center;gap:16px}
header h1{font-size:16px;font-weight:600}
header p{font-size:12px;color:#64748b;margin-top:2px}
.warn{background:#451a03;border:1px solid #92400e;border-radius:8px;padding:10px 14px;margin:16px 24px;font-size:12px;color:#fbbf24;display:flex;gap:8px;align-items:flex-start}
.main{padding:16px 24px;display:grid;grid-template-columns:1fr 320px;gap:16px;max-width:1400px}
@media(max-width:900px){.main{grid-template-columns:1fr}}
.editor-wrap{background:#1e2433;border-radius:10px;border:1px solid #2d3748;overflow:hidden}
.editor-toolbar{padding:10px 14px;border-bottom:1px solid #2d3748;display:flex;gap:8px;align-items:center;flex-wrap:wrap}
.editor-toolbar span{font-size:12px;color:#64748b;flex:1}
button{padding:6px 14px;border-radius:6px;border:1px solid #2d3748;background:#161b27;color:#94a3b8;font-size:12px;font-family:inherit;cursor:pointer;transition:all .15s}
button:hover{border-color:#4b5563;color:#e2e8f0}
button.primary{background:#1d4ed8;border-color:#1d4ed8;color:white}
button.primary:hover{background:#1e40af}
button.danger{background:#7f1d1d;border-color:#7f1d1d;color:#fca5a5}
button.danger:hover{background:#991b1b}
button.success{background:#14532d;border-color:#14532d;color:#86efac}
button.success:hover{background:#166534}
textarea{width:100%;min-height:520px;background:#0d1117;color:#e2e8f0;border:none;padding:16px;font-family:\'IBM Plex Mono\',monospace;font-size:13px;line-height:1.6;resize:vertical;outline:none;tab-size:2}
.panel{display:flex;flex-direction:column;gap:12px}
.card{background:#1e2433;border-radius:10px;border:1px solid #2d3748;padding:16px}
.card h3{font-size:12px;font-weight:600;color:#94a3b8;text-transform:uppercase;letter-spacing:.5px;margin-bottom:12px}
.status{padding:10px 12px;border-radius:6px;font-size:12px;font-family:\'IBM Plex Mono\',monospace;min-height:44px;line-height:1.5;white-space:pre-wrap;word-break:break-all}
.status.ok{background:#052e16;color:#86efac;border:1px solid #14532d}
.status.err{background:#450a0a;color:#fca5a5;border:1px solid #7f1d1d}
.status.idle{background:#161b27;color:#475569;border:1px solid #2d3748}
.field{margin-bottom:10px}
.field label{display:block;font-size:11px;color:#64748b;margin-bottom:4px}
.field input,.field select{width:100%;padding:7px 10px;background:#0d1117;border:1px solid #2d3748;border-radius:6px;color:#e2e8f0;font-size:13px;font-family:inherit}
.field input:focus,.field select:focus{outline:none;border-color:#3b82f6}
.busqueda-item{padding:10px 12px;background:#161b27;border-radius:6px;border:1px solid #2d3748;margin-bottom:8px;cursor:pointer;transition:border-color .15s}
.busqueda-item:hover{border-color:#3b82f6}
.busqueda-item .bi-name{font-size:13px;font-weight:500}
.busqueda-item .bi-meta{font-size:11px;color:#64748b;margin-top:3px}
.dot{width:8px;height:8px;border-radius:50%;display:inline-block;margin-right:6px}
.dot-on{background:#22c55e}.dot-off{background:#475569}
hr{border:none;border-top:1px solid #2d3748;margin:8px 0}
.token-note{font-size:11px;color:#475569;margin-top:8px;line-height:1.5}
</style>
</head>
<body>
<header>
  <div>
    <h1>⚙️ Editar configuración</h1>
    <p>AutoScout24 Monitor — config.json</p>
  </div>
</header>
<div class="warn">
  ⚠️ <span>Esta página no está enlazada desde ningún otro lugar del dashboard. Solo es accesible conociendo su URL directa. Los cambios se guardan directamente en el repositorio de GitHub vía API.</span>
</div>
<div class="main">
  <div class="editor-wrap">
    <div class="editor-toolbar">
      <span id="lbl-editor">Cargando config.json...</span>
      <button onclick="formatear()">Formatear JSON</button>
      <button onclick="validar()">✓ Validar</button>
      <button class="success" onclick="guardar()">💾 Guardar en GitHub</button>
    </div>
    <textarea id="editor" spellcheck="false" oninput="onEdit()"></textarea>
  </div>

  <div class="panel">
    <div class="card">
      <h3>Estado validación</h3>
      <div class="status idle" id="status-val">Pendiente de validar</div>
    </div>

    <div class="card">
      <h3>Guardar en GitHub</h3>
      <div class="field">
        <label>GitHub Token (PAT con permiso contents:write)</label>
        <input type="password" id="gh-token" placeholder="ghp_..." autocomplete="off">
      </div>
      <div class="field">
        <label>Repositorio (owner/repo)</label>
        <input type="text" id="gh-repo" placeholder="antrivasclaude-glitch/autoscout-monitor">
      </div>
      <div class="field">
        <label>Branch</label>
        <select id="gh-branch">
          <option value="main">main</option>
          <option value="master">master</option>
        </select>
      </div>
      <button class="primary" style="width:100%" onclick="guardar()">💾 Guardar config.json</button>
      <div class="status idle" id="status-save" style="margin-top:10px">Sin cambios guardados</div>
      <p class="token-note">El token se usa solo en esta sesión del navegador y no se almacena en ningún servidor.</p>
    </div>

    <div class="card">
      <h3>Búsquedas detectadas</h3>
      <div id="busquedas-list"><p style="font-size:12px;color:#475569">Carga el JSON para ver el resumen</p></div>
    </div>

    <div class="card">
      <h3>Referencia de campos</h3>
      <div style="font-size:11px;color:#64748b;line-height:1.8">
        <b style="color:#93c5fd">activa</b>: true / false<br>
        <b style="color:#93c5fd">fuel</b>: "D" diésel · "B" gasolina · "E" eléctrico · null todos<br>
        <b style="color:#93c5fd">pais</b>: es · de · fr · it · nl · be · at · lu<br>
        <b style="color:#93c5fd">precio_max</b>: número en € (null = sin límite)<br>
        <b style="color:#93c5fd">km_max</b>: número en km (null = sin límite)<br>
        <b style="color:#93c5fd">anio_min/max</b>: año con 4 dígitos (null = sin límite)<br>
        <b style="color:#93c5fd">max_paginas</b>: 1–20 (20 anuncios por página)<br>
        <b style="color:#93c5fd">adjuntar_hoja_calculo</b>: true / false
      </div>
    </div>
  </div>
</div>

<script>
const REPO_DEFAULT = "antrivasclaude-glitch/autoscout-monitor";

// ── Cargar config.json desde gh-pages o GitHub API ────────────────────────
async function cargarConfig() {
  const lbl = document.getElementById("lbl-editor");
  // Intentar cargar el raw de main branch via GitHub API (sin auth, si es público)
  const repo = document.getElementById("gh-repo").value || REPO_DEFAULT;
  lbl.textContent = "Cargando config.json desde GitHub...";
  try {
    const url = `https://api.github.com/repos/${repo}/contents/config.json`;
    const r = await fetch(url, {headers:{"Accept":"application/vnd.github+json"}});
    if (r.ok) {
      const data = await r.json();
      const decoded = atob(data.content.replace(/\n/g,""));
      document.getElementById("editor").value = JSON.stringify(JSON.parse(decoded), null, 2);
      lbl.textContent = "config.json cargado desde GitHub ✓";
      actualizarResumen();
      return;
    }
  } catch(e) {}
  lbl.textContent = "No se pudo cargar automáticamente — pega el JSON manualmente";
}

window.addEventListener("load", () => {
  const saved = localStorage.getItem("as24-gh-repo");
  if (saved) document.getElementById("gh-repo").value = saved;
  const branch = localStorage.getItem("as24-gh-branch");
  if (branch) document.getElementById("gh-branch").value = branch;
  cargarConfig();
});

document.getElementById("gh-repo").addEventListener("change", e => {
  localStorage.setItem("as24-gh-repo", e.target.value);
});
document.getElementById("gh-branch").addEventListener("change", e => {
  localStorage.setItem("as24-gh-branch", e.target.value);
});

function onEdit() {
  document.getElementById("status-val").className = "status idle";
  document.getElementById("status-val").textContent = "Modificado — valida antes de guardar";
  actualizarResumen();
}

function formatear() {
  try {
    const v = JSON.parse(document.getElementById("editor").value);
    document.getElementById("editor").value = JSON.stringify(v, null, 2);
    validar();
  } catch(e) {
    setStatus("status-val", "err", "JSON inválido:\n" + e.message);
  }
}

function validar() {
  const txt = document.getElementById("editor").value.trim();
  if (!txt) { setStatus("status-val","err","El editor está vacío"); return false; }
  try {
    const cfg = JSON.parse(txt);
    const bs = cfg.busquedas;
    if (!Array.isArray(bs)) { setStatus("status-val","err","Falta el array \"busquedas\""); return false; }
    const errores = [];
    bs.forEach((b,i) => {
      if (!b.nombre) errores.push(`busquedas[${i}]: falta "nombre"`);
      if (!b.marca)  errores.push(`busquedas[${i}]: falta "marca"`);
      if (!b.modelo) errores.push(`busquedas[${i}]: falta "modelo"`);
      if (b.max_paginas && (b.max_paginas < 1 || b.max_paginas > 20))
        errores.push(`busquedas[${i}]: max_paginas debe estar entre 1 y 20`);
    });
    if (errores.length) { setStatus("status-val","err", errores.join("\n")); return false; }
    setStatus("status-val","ok",`✓ JSON válido\n${bs.length} búsquedas · ${bs.filter(b=>b.activa).length} activas`);
    return true;
  } catch(e) {
    setStatus("status-val","err","JSON inválido:\n" + e.message);
    return false;
  }
}

async function guardar() {
  if (!validar()) return;
  const token = document.getElementById("gh-token").value.trim();
  const repo  = document.getElementById("gh-repo").value.trim() || REPO_DEFAULT;
  const branch= document.getElementById("gh-branch").value;
  if (!token) { setStatus("status-save","err","Introduce el GitHub Token (PAT)"); return; }

  setStatus("status-save","idle","Guardando...");

  try {
    // Obtener SHA actual del archivo
    const urlFile = `https://api.github.com/repos/${repo}/contents/config.json`;
    const rGet = await fetch(urlFile, {
      headers:{"Authorization":`Bearer ${token}`,"Accept":"application/vnd.github+json"}
    });
    let sha = null;
    if (rGet.ok) { sha = (await rGet.json()).sha; }

    const content = btoa(unescape(encodeURIComponent(document.getElementById("editor").value)));
    const payload = {
      message: "[config] Actualizar config.json via editor web",
      content,
      branch,
    };
    if (sha) payload.sha = sha;

    const rPut = await fetch(urlFile, {
      method: "PUT",
      headers:{
        "Authorization": `Bearer ${token}`,
        "Accept": "application/vnd.github+json",
        "Content-Type": "application/json"
      },
      body: JSON.stringify(payload)
    });
    if (rPut.ok) {
      setStatus("status-save","ok",`✓ config.json guardado en ${repo} (${branch})\n${new Date().toLocaleString("es-ES")}`);
    } else {
      const err = await rPut.json();
      setStatus("status-save","err","Error " + rPut.status + ":\n" + (err.message||JSON.stringify(err)));
    }
  } catch(e) {
    setStatus("status-save","err","Error de red:\n" + e.message);
  }
}

function setStatus(id, type, msg) {
  const el = document.getElementById(id);
  el.className = "status " + type;
  el.textContent = msg;
}

function actualizarResumen() {
  try {
    const cfg = JSON.parse(document.getElementById("editor").value);
    const bs = cfg.busquedas || [];
    document.getElementById("busquedas-list").innerHTML = bs.map(b => `
      <div class="busqueda-item">
        <div class="bi-name">
          <span class="dot ${b.activa ? \'dot-on\' : \'dot-off\'}"></span>${b.nombre || \'(sin nombre)\'}
        </div>
        <div class="bi-meta">${b.marca||\'\'} ${b.modelo||\'\'} · ${b.pais||\'es\'} · ${b.fuel||\'todos\'} · hasta ${b.precio_max ? b.precio_max.toLocaleString(\'es-ES\')+\'€\' : \'—\'}</div>
      </div>`).join("") || \'<p style="font-size:12px;color:#475569">Sin búsquedas</p>\';
  } catch(e) {}
}
</script>
</body>
</html>'''


def _gh_get_sha(base: str, path: str, hdrs: dict, ref: str = "gh-pages") -> str | None:
    """Devuelve el SHA de un archivo en gh-pages o None si no existe."""
    r = requests.get(f"{base}/contents/{path}", headers=hdrs,
                     params={"ref": ref}, timeout=15)
    return r.json().get("sha") if r.status_code == 200 else None


def _gh_put(base: str, path: str, content_bytes: bytes, hdrs: dict,
            message: str, sha: str | None = None) -> bool:
    """Crea o actualiza un archivo en gh-pages. Devuelve True si OK."""
    payload: dict = {
        "message": message,
        "content": base64.b64encode(content_bytes).decode(),
        "branch": "gh-pages",
    }
    if sha:
        payload["sha"] = sha
    r = requests.put(f"{base}/contents/{path}", headers=hdrs,
                     json=payload, timeout=30)
    return r.status_code in (200, 201)


def _gh_delete(base: str, path: str, sha: str, hdrs: dict) -> bool:
    """Borra un archivo de gh-pages. Devuelve True si OK."""
    r = requests.delete(f"{base}/contents/{path}", headers=hdrs,
                        json={"message": f"[log] Borrar {path} (>7 días) [skip ci]",
                              "sha": sha, "branch": "gh-pages"}, timeout=15)
    return r.status_code == 200


def publicar_github_pages(paginas: dict) -> str | None:
    """
    Publica archivos HTML + log diario en gh-pages y borra logs >7 días.
    Requiere GITHUB_TOKEN y GITHUB_REPOSITORY en el entorno.
    """
    token = os.environ.get("GITHUB_TOKEN", "")
    repo  = os.environ.get("GITHUB_REPOSITORY", "")

    if not token or not repo:
        log.warning("Sin GITHUB_TOKEN o GITHUB_REPOSITORY — GitHub Pages omitido")
        return None

    hdrs = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    base = f"https://api.github.com/repos/{repo}"

    # Asegurar que la rama gh-pages existe
    r = requests.get(f"{base}/branches/gh-pages", headers=hdrs, timeout=15)
    if r.status_code == 404:
        log.info("Creando rama gh-pages...")
        r2 = requests.get(f"{base}", headers=hdrs, timeout=15)
        default = r2.json().get("default_branch", "main")
        r3 = requests.get(f"{base}/git/ref/heads/{default}", headers=hdrs, timeout=15)
        if r3.status_code == 200:
            sha = r3.json()["object"]["sha"]
            requests.post(f"{base}/git/refs", headers=hdrs,
                          json={"ref": "refs/heads/gh-pages", "sha": sha}, timeout=15)
        else:
            log.error("No se pudo crear gh-pages")
            return None

    # ── Publicar archivos HTML ────────────────────────────────
    ok = 0
    for path, html_content in paginas.items():
        sha_file = _gh_get_sha(base, path, hdrs)
        if _gh_put(base, path, html_content.encode("utf-8"), hdrs,
                   f"[dashboard] {path} [skip ci]", sha_file):
            ok += 1
            log.debug(f"gh-pages: {path} publicado")
        else:
            log.error(f"gh-pages error en {path}")

    log.info(f"GitHub Pages: {ok}/{len(paginas)} archivos HTML publicados")

    # ── Log diario: publicar log_YYYY-MM-DD.txt ───────────────
    hoy_str = date.today().isoformat()
    log_path = f"log_{hoy_str}.txt"
    log_content = "\n".join(_log_lines).encode("utf-8")
    sha_log = _gh_get_sha(base, log_path, hdrs)
    if _gh_put(base, log_path, log_content, hdrs,
               f"[log] {log_path} [skip ci]", sha_log):
        log.info(f"Log publicado: {log_path}")
    else:
        log.warning(f"No se pudo publicar el log {log_path}")

    # ── Borrar logs de más de 7 días ──────────────────────────
    from datetime import timedelta
    limite = date.today() - timedelta(days=7)
    try:
        r = requests.get(f"{base}/git/trees/gh-pages", headers=hdrs,
                         params={"recursive": "1"}, timeout=15)
        if r.status_code == 200:
            archivos = [f["path"] for f in r.json().get("tree", [])
                        if f["path"].startswith("log_") and f["path"].endswith(".txt")]
            for ap in archivos:
                # log_YYYY-MM-DD.txt → extraer fecha
                m = re.search(r"log_(\d{4}-\d{2}-\d{2})\.txt", ap)
                if m:
                    try:
                        fecha_log = date.fromisoformat(m.group(1))
                        if fecha_log < limite:
                            sha_old = _gh_get_sha(base, ap, hdrs)
                            if sha_old and _gh_delete(base, ap, sha_old, hdrs):
                                log.info(f"Log antiguo borrado: {ap}")
                    except ValueError:
                        pass
    except Exception as e:
        log.warning(f"Error al limpiar logs antiguos: {e}")

    owner, rname = repo.split("/", 1)
    url_pages = f"https://{owner}.github.io/{rname}/"
    log.info(f"Dashboard URL: {url_pages}")
    return url_pages


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

def main():
    inicio = datetime.now()
    sep    = "═" * 60
    log.info(sep)
    log.info(f"AutoScout24 Agent — INICIO {inicio.strftime('%Y-%m-%d %H:%M:%S')}")
    log.info(sep)

    busquedas = [b for b in CFG.get("busquedas", []) if b.get("activa", True)]
    if not busquedas:
        log.warning("No hay búsquedas activas en config.json")
        return

    log.info(f"Búsquedas activas: {len(busquedas)}")

    resultados     = []
    url_sheets     = ""
    paginas_html   = {}   # {path: html_content} para GitHub Pages
    busquedas_info = []   # metadata para index.html

    try:
        for i, b in enumerate(busquedas, 1):
            nombre = b.get("nombre", f"Búsqueda {i}")
            log.info(sep)
            log.info(f"BÚSQUEDA {i}/{len(busquedas)}: {nombre}")
            log.info(sep)

            # 1. Estado anterior
            estado_anterior = cargar_estado(nombre)

            # 2. Scraping
            anuncios = scrape_busqueda(b)
            if not anuncios:
                log.warning(f"[{nombre}] Sin anuncios — posible bloqueo temporal")
                resultados.append({"nombre": nombre, "nuevos": [], "bajadas": [], "url_sheets": ""})
                continue

            # 3. Detectar cambios
            nuevos, bajadas = detectar_cambios(anuncios, estado_anterior)

            # 4. Google Sheets (actualiza los datos)
            url_sheets = actualizar_sheets_busqueda(nombre, anuncios, nuevos, bajadas) or url_sheets

            # 5. Guardar estado compacto
            guardar_estado(nombre, anuncios)

            # 6. Leer histórico completo de Sheets
            filas_hist = leer_hoja_completa(nombre)
            
            # 7. Verificar anuncios activos (solo filtra la lista, no modifica Sheets todavía)
            log.info(f"[{nombre}] Verificando disponibilidad de anuncios...")
            filas_hist_limpias = verificar_anuncios_activos(filas_hist)
            
            # 7b. Limpieza de Sheets cada lunes (elimina anuncios no disponibles)
            es_lunes = datetime.now().weekday() == 0
            if es_lunes and len(filas_hist_limpias) < len(filas_hist):
                log.info(f"[{nombre}] Limpieza automática (lunes): eliminando {len(filas_hist) - len(filas_hist_limpias)} anuncios no disponibles de Sheets")
                try:
                    sp = conectar_sheets()
                    ws = sp.worksheet(nombre[:50])
                    ws.clear()
                    if filas_hist_limpias:
                        # Reconstruir filas para Sheets
                        filas_sheets = [CABECERAS]
                        for f in filas_hist_limpias:
                            fila_datos = [
                                f.get("id",""), f.get("titulo",""), f.get("precio",""),
                                f.get("precio_anterior",""), "", "",
                                f.get("anio",""), f.get("km",""), f.get("combustible",""),
                                f.get("ubicacion",""), f.get("estado",""), 
                                f.get("fecha_detectado",""), f.get("url","")
                            ]
                            filas_sheets.append(fila_datos)
                        ws.update(filas_sheets, "A1")
                        log.info(f"[{nombre}] Sheets limpiado: {len(filas_hist) - len(filas_hist_limpias)} anuncios eliminados")
                except Exception as e:
                    log.error(f"[{nombre}] Error limpiando Sheets: {e}")
            elif es_lunes:
                log.info(f"[{nombre}] Limpieza lunes: todos los anuncios están activos")
            
            # Usar anuncios activos para email y dashboard
            filas_hist = filas_hist_limpias

            # 8. Email independiente por búsqueda + adjuntos (con filas activas)
            adjuntar_hoja = b.get("adjuntar_hoja_calculo", False)
            enviar_email_busqueda(
                nombre, nuevos, bajadas, url_sheets,
                filas_hist=filas_hist, adjuntar_hoja=adjuntar_hoja
            )

            # 10. Preparar datos para el dashboard HTML
            slug = _slug(nombre)
            paginas_html[f"busquedas/{slug}.html"] = generar_html_busqueda(
                nombre, filas_hist, nuevos, bajadas, date.today().isoformat()
            )
            precios = [int(f["precio"]) for f in filas_hist if f.get("precio") and str(f["precio"]).isdigit()]
            busquedas_info.append({
                "nombre":     nombre,
                "slug":       slug,
                "total":      len(filas_hist),
                "nuevos":     len(nuevos),
                "bajadas":    len(bajadas),
                "min_precio": f"{min(precios):,} €".replace(",", ".") if precios else "—",
                "med_precio": f"{sum(precios)//len(precios):,} €".replace(",", ".") if precios else "—",
                "fecha":      date.today().isoformat(),
            })

            resultados.append({"nombre": nombre, "nuevos": nuevos, "bajadas": bajadas, "url_sheets": url_sheets})

        # 9. Generar index.html + editor configuracion y publicar en GitHub Pages
        if busquedas_info:
            log.info(sep)
            log.info("Generando y publicando dashboard en GitHub Pages...")
            paginas_html["index.html"] = generar_index_html(busquedas_info)
            # Editor config y ver_registros: no enlazados, solo accesibles por URL directa
            paginas_html["editar_configuracion.html"] = generar_editor_configuracion()
            paginas_html["ver_registros.html"] = generar_ver_registros()
            url_pages = publicar_github_pages(paginas_html)
            if url_pages:
                log.info(f"Dashboard: {url_pages}")
                log.info(f"Editor config: {url_pages}editar_configuracion.html")

    except Exception as e:
        log.critical(f"Error crítico: {e}", exc_info=True)
        raise
    finally:
        limpiar_temporales()
        duracion = (datetime.now() - inicio).total_seconds()
        log.info(sep)
        log.info(f"AutoScout24 Agent — FIN | Duración: {duracion:.1f}s")
        log.info(sep)


if __name__ == "__main__":
    main()
    
